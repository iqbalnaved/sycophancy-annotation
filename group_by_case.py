import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

xl = pd.ExcelFile('/mnt/d/Naved/analysis/helpmed/llm-delusions-annotations_helpmed_gpt-5.5-2026-04-23.xlsx')
runs = ['run0', 'run1', 'run2']

header_colors = {'run0': '2F5496', 'run1': '375623', 'run2': '7B2C2C'}
alt_colors    = {'run0': 'DCE6F1', 'run1': 'EBF1DE', 'run2': 'F2DCDB'}

wb = Workbook()
ws = wb.active
ws.title = "Grouped by case_id"

headers = ['case_id', 'total_positives', 'codes_flagged', 'groups_flagged', 'scores']
col_widths = [14, 16, 55, 28, 20]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

current_row = 1

for run in runs:
    df = xl.parse(run)
    pos = df[df['binary_label'] == 1]
    grouped = pos.groupby('case_id').agg(
        total_positives=('binary_label', 'count'),
        codes_flagged=('code', lambda x: ', '.join(sorted(x.unique()))),
        groups_flagged=('group', lambda x: ', '.join(sorted(x.unique()))),
        scores=('score', lambda x: ', '.join(str(v) for v in sorted(x.tolist())))
    ).reset_index()

    # Section label
    label_cell = ws.cell(row=current_row, column=1, value=run.upper())
    label_cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    label_cell.fill = PatternFill('solid', start_color='1F1F1F')
    label_cell.alignment = Alignment(horizontal='left')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    current_row += 1

    # Header row
    hfill = PatternFill('solid', start_color=header_colors[run])
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=current_row, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='top')
    current_row += 1

    # Data rows
    afill = PatternFill('solid', start_color=alt_colors[run])
    for i, row in enumerate(grouped.itertuples(index=False)):
        fill = afill if i % 2 == 0 else None
        for col, val in enumerate(row, 1):
            c = ws.cell(row=current_row, column=col, value=val)
            c.font = Font(name='Arial', size=10)
            c.alignment = Alignment(horizontal='left' if col > 2 else 'center',
                                    vertical='top', wrap_text=True)
            if fill:
                c.fill = fill
        current_row += 1

    current_row += 2

wb.save('/mnt/d/Naved/analysis/helpmed/grouped_by_case_id_all_runs.xlsx')