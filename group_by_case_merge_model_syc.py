import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Config ────────────────────────────────────────────────────────────────────
# ANNOTATIONS_CSV = "/mnt/d/Naved/Outputs/helpmed/llm-delusions-annotations_helpmed_gpt-5.5-2026-04-23_run0_2026-05-22-10-51-01.csv"
# ANNOTATIONS_CSV = "/mnt/d/Naved/Outputs/helpmed/llm-delusions-annotations_helpmed_gpt-5.5-2026-04-23_run1_2026-05-22-10-55-13.csv"
ANNOTATIONS_CSV = "/mnt/d/Naved/Outputs/helpmed/llm-delusions-annotations_helpmed_gpt-5.5-2026-04-23_run2_2026-05-22-10-55-17.csv"
SYC_CSV         = '/mnt/d/Naved/Outputs/helpmed/sycophancy_gpt-5.2-2025-12-11_02262026_1210.csv'

treatments_map = {1: 'Llama 3 70B', 2: 'GPT-4o', 3: 'Control', 4: 'Command R+'}

header_colors = {'run0': '2F5496', 'run1': '375623', 'run2': '7B2C2C'}
alt_colors    = {'run0': 'DCE6F1', 'run1': 'EBF1DE', 'run2': 'F2DCDB'}

# ── Infer RUN from filename ───────────────────────────────────────────────────
match = re.search(r'(run\d+)', ANNOTATIONS_CSV)
assert match, "Could not infer run from filename — expected 'run0', 'run1', etc."
RUN = match.group(1)
timestamp  = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

OUTPUT = f'/mnt/d/Naved/analysis/helpmed/annotation_gpt5.5_sycophancy_gpt5.2_{RUN}_{timestamp}.xlsx'

# ── Load ──────────────────────────────────────────────────────────────────────
df  = pd.read_csv(ANNOTATIONS_CSV)
syc = pd.read_csv(SYC_CSV)
ex  = pd.read_csv('/mnt/d/Naved/Codes/HELPMed/data/main/clean_examples.csv')

# Get model from ex
ex['model'] = ex['treatment_id'].map(treatments_map)
id_to_model = ex[['id', 'model']].drop_duplicates('id').rename(columns={'id': 'case_id'})


# ── Group ─────────────────────────────────────────────────────────────────────
pos = df[df['binary_label'] == 1]
grouped = pos.groupby('case_id').agg(
    total_positives=('binary_label', 'count'),
    codes_flagged=('code',  lambda x: ', '.join(sorted(x.unique()))),
    groups_flagged=('group', lambda x: ', '.join(sorted(x.unique()))),
    scores=('score', lambda x: ', '.join(str(v) for v in sorted(x.tolist())))
).reset_index()

# ── Merge ─────────────────────────────────────────────────────────────────────
syc_slim = syc[['id', 'is_sycophantic']].rename(columns={'id': 'case_id'})

grouped = grouped.merge(syc_slim, on='case_id', how='left')
grouped = grouped.merge(id_to_model, on='case_id', how='left')

cols = ['case_id', 'model', 'is_sycophantic', 'total_positives', 'codes_flagged', 'groups_flagged', 'scores']
grouped = grouped[cols]

# ── Write Excel ───────────────────────────────────────────────────────────────
col_widths = [14, 16, 16, 16, 55, 28, 20]

wb = Workbook()
ws = wb.active
ws.title = f"Grouped by case_id ({RUN})"

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

current_row = 1

label_cell = ws.cell(row=current_row, column=1, value=RUN.upper())
label_cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
label_cell.fill = PatternFill('solid', start_color='1F1F1F')
label_cell.alignment = Alignment(horizontal='left')
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(cols))
current_row += 1

hfill = PatternFill('solid', start_color=header_colors.get(RUN, '404040'))
hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
for col, h in enumerate(cols, 1):
    c = ws.cell(row=current_row, column=col, value=h)
    c.font = hfont
    c.fill = hfill
    c.alignment = Alignment(horizontal='center', vertical='top')
current_row += 1

afill = PatternFill('solid', start_color=alt_colors.get(RUN, 'E0E0E0'))
for i, row in enumerate(grouped.itertuples(index=False)):
    fill = afill if i % 2 == 0 else None
    for col, val in enumerate(row, 1):
        c = ws.cell(row=current_row, column=col, value=val)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(
            horizontal='left' if col > 2 else 'center',
            vertical='top', wrap_text=True
        )
        if fill:
            c.fill = fill
    current_row += 1

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")